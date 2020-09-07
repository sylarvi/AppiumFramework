# -*-coding:utf-8 -*-
# @Author: lixiao
# Created on: 2020-06-15

import openpyxl
from Config.project_var import excel_path


class HandleExcel:
    def load_excel(self):
        """加载excel文件"""
        excel_obj = openpyxl.load_workbook(excel_path)
        return excel_obj

    def get_sheet(self, index=None):
        """加载sheet信息"""
        sheet_name = self.load_excel().sheetnames
        if index is None:
            index = 0
        sheet_data = self.load_excel()[sheet_name[index]]
        return sheet_data

    def get_cell_data(self, row, cols):
        """获取某一个单元格的数据"""
        data = self.get_sheet().cell(row=row, column=cols).value
        return data

    def get_rows_number(self):
        """获取总行数"""
        rows = self.get_sheet().max_row
        return rows

    def get_row_data(self, row):
        """获取某一行的数据"""
        data_list = []
        for cell in self.get_sheet()[row]:
            data_list.append(cell.value)
        return data_list

    def write_data(self, row, cols, data):
        """向某一单元格写入数据"""
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row, cols, data)
        wb.save(excel_path)

    def get_column_data(self, key=None):
        """获取某一列的数据"""
        data_list = []
        if key is None:
            key = 'A'
        columns_data_list = self.get_sheet()[key]
        for cell in columns_data_list:
            data_list.append(cell.value)
        return data_list

    def get_case_number(self, case_id):
        """获取某一条case所在的行号"""
        mark = 1
        cols_data = self.get_column_data()
        for data in cols_data:
            if case_id == data:
                return mark
            mark += 1
        return mark

    def get_all_data(self):
        """获取excel中所有的数据"""
        data_list = []
        for cell in range(1, self.get_rows_number()):
            data_list.append(self.get_row_data(cell + 1))
        return data_list


if __name__ == '__main__':
    exc = HandleExcel()
    print(type(exc.get_cell_data(2, 4)))
    str = exc.get_cell_data(2, 4)
    list = str.split('\n')
    print(list)
    print(exc.get_case_number('case02'))
